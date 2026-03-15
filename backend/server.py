"""InvoiceAI Backend — FastAPI + MongoDB + Claude Vision AI + Cost Optimizations"""

from contextlib import asynccontextmanager
from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Query, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import JSONResponse
from motor.motor_asyncio import AsyncIOMotorClient
import os
import sys
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any, Annotated
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
from jose import JWTError, jwt
import re
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.middleware import SlowAPIMiddleware
from slowapi.errors import RateLimitExceeded
from llm_providers import LLMMessage, build_manager_from_env
import json
import base64
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv

# Payment utilities
from payment_utils import (
    create_razorpay_order,
    verify_razorpay_payment,
    verify_razorpay_webhook,
    rupees_to_paise,
    paise_to_rupees,
    is_razorpay_configured,
    get_configured_gateways
)

# Sprint 1 cost-reduction services
sys.path.insert(0, str(Path(__file__).parent))
from services.image_preprocessor import preprocess_invoice_image, get_image_hash
from services.model_router import select_model, should_use_batch, HAIKU, SONNET
from services.pdf_handler import process_pdf
from services.batch_processor import (
    queue_for_batch, process_batch_queue, extract_invoice_with_llm,
    EXTRACTION_SYSTEM_PROMPT, setup_scheduler
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

_llm_manager = None


def _get_llm_manager():
    global _llm_manager
    if _llm_manager is None:
        _llm_manager = build_manager_from_env()
    return _llm_manager

# ─── Config ───────────────────────────────────────────────────────────────────
mongo_url = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
db_name = os.environ.get("DB_NAME", "invoiceai_db")
JWT_SECRET = os.environ.get("JWT_SECRET", "")
JWT_ALGORITHM = "HS256"
JWT_EXPIRE_DAYS = 30

BATCH_ENABLED = True  # Feature flag for batch processing

ALLOWED_MIME_TYPES = {"image/jpeg", "image/png", "application/pdf"}
MAX_UPLOAD_BYTES = int(os.environ.get("MAX_UPLOAD_BYTES", 10 * 1024 * 1024))
EXPORT_MAX_RECORDS = int(os.environ.get("EXPORT_MAX_RECORDS", 5000))

CORS_ALLOW_ORIGINS = [
    origin.strip()
    for origin in os.environ.get(
        "CORS_ALLOW_ORIGINS",
        "http://localhost:8081,http://localhost:19006"
    ).split(",")
    if origin.strip()
]

RATE_LIMIT_DEFAULT = os.environ.get("RATE_LIMIT_DEFAULT", "60/minute")
RATE_LIMIT_AUTH = os.environ.get("RATE_LIMIT_AUTH", "5/minute")
RATE_LIMIT_PROCESS = os.environ.get("RATE_LIMIT_PROCESS", "10/minute")
RATE_LIMIT_EXPORT = os.environ.get("RATE_LIMIT_EXPORT", "10/minute")
REDIS_URL = os.environ.get("REDIS_URL", "redis://localhost:6379/0")

if not JWT_SECRET or len(JWT_SECRET) < 32:
    raise RuntimeError("JWT_SECRET must be set and at least 32 characters long")

mongo_client = AsyncIOMotorClient(mongo_url)
db = mongo_client[db_name]

@asynccontextmanager
async def lifespan(app: FastAPI):
    setup_scheduler(app, db)
    yield
    if hasattr(app.state, 'scheduler'):
        app.state.scheduler.shutdown(wait=False)
    mongo_client.close()

def _rate_limit_key(request: Request) -> str:
    user_id = getattr(request.state, "user_id", None)
    return user_id or get_remote_address(request)


limiter = Limiter(key_func=_rate_limit_key, default_limits=[RATE_LIMIT_DEFAULT], storage_uri=REDIS_URL)


class MaxRequestSizeMiddleware(BaseHTTPMiddleware):
    def __init__(self, app: FastAPI, max_body_size: int) -> None:
        super().__init__(app)
        self._max_body_size = max_body_size

    async def dispatch(self, request: Request, call_next):
        content_length = request.headers.get("content-length")
        if content_length and int(content_length) > self._max_body_size:
            return JSONResponse(
                status_code=413,
                content={"detail": "Request payload too large"}
            )
        return await call_next(request)


class UserIdContextMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        auth_header = request.headers.get("authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header.split(" ", 1)[1]
            try:
                payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
                request.state.user_id = payload.get("sub")
            except JWTError:
                pass
        return await call_next(request)


app = FastAPI(title="InvoiceAI API", lifespan=lifespan)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(name)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# ─── Pydantic Models ──────────────────────────────────────────────────────────

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    business_name: str
    gstin: Optional[str] = None
    business_type: Optional[str] = "retail"
    push_token: Optional[str] = None  # Expo push notification token

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserOut(BaseModel):
    id: str
    email: str
    name: str
    business_name: str
    gstin: Optional[str] = None
    business_type: Optional[str] = None
    plan: str = "free"
    credits: int = 100
    monthly_invoice_count: int = 0
    monthly_invoice_limit: Optional[int] = 10
    created_at: datetime

class LineItemModel(BaseModel):
    description: str
    hsn_code: Optional[str] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    unit_price: Optional[float] = None
    line_total: float = 0.0
    gst_rate: Optional[float] = None

class InvoiceCreate(BaseModel):
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    due_date: Optional[str] = None
    vendor_name: Optional[str] = None
    vendor_gstin: Optional[str] = None
    vendor_phone: Optional[str] = None
    vendor_address: Optional[str] = None
    buyer_name: Optional[str] = None
    buyer_gstin: Optional[str] = None
    line_items: List[LineItemModel] = []
    subtotal: Optional[float] = None
    discount: Optional[float] = 0.0
    cgst: Optional[float] = 0.0
    sgst: Optional[float] = 0.0
    igst: Optional[float] = 0.0
    total_tax: Optional[float] = 0.0
    grand_total: float = 0.0
    currency: str = "INR"
    payment_terms: Optional[str] = None
    status: str = "pending"
    category: Optional[str] = None
    notes: Optional[str] = None
    source_type: str = "camera"
    confidence_score: Optional[int] = None
    is_duplicate: bool = False
    has_anomaly: bool = False

class InvoiceOut(InvoiceCreate):
    id: str
    user_id: str
    created_at: datetime
    updated_at: datetime

class ProcessInvoiceRequest(BaseModel):
    image_base64: str
    source_type: str = "camera"
    mime_type: str = "image/jpeg"

class ExportRequest(BaseModel):
    format: str = "excel"
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    vendor_name: Optional[str] = None
    status_filter: Optional[str] = None

class CorrectionCreate(BaseModel):
    invoice_id: str
    field_name: str
    original_value: str
    corrected_value: str

class BulkUpdateRequest(BaseModel):
    invoice_ids: List[str]
    action: str  # mark_paid | delete | categorize
    updates: Optional[Dict[str, Any]] = None

class PushTokenUpdate(BaseModel):
    push_token: str

class InvoiceStatusUpdate(BaseModel):
    status: str

# ─── Subscription & Credits Models ───────────────────────────────────────────

class SubscriptionTier(str):
    FREE = "free"
    STARTER = "starter"
    PRO = "pro"

class PlanConfig(BaseModel):
    tier: str
    name: str
    price_monthly: float
    currency: str = "INR"
    monthly_invoice_limit: Optional[int]
    api_calls_per_hour: int
    initial_credits: int
    features: List[str]

# Plan definitions with security limits
PLAN_CONFIGS = {
    "free": PlanConfig(
        tier="free",
        name="Free",
        price_monthly=0.0,
        monthly_invoice_limit=10,
        api_calls_per_hour=10,
        initial_credits=100,
        features=["10 invoices/month", "Basic OCR", "CSV export", "Email support"]
    ),
    "starter": PlanConfig(
        tier="starter",
        name="Starter",
        price_monthly=299.0,
        monthly_invoice_limit=100,
        api_calls_per_hour=60,
        initial_credits=1000,
        features=["100 invoices/month", "Advanced OCR", "Excel & CSV export", "Priority support", "Batch processing"]
    ),
    "pro": PlanConfig(
        tier="pro",
        name="Pro",
        price_monthly=999.0,
        monthly_invoice_limit=None,  # Unlimited
        api_calls_per_hour=300,
        initial_credits=5000,
        features=["Unlimited invoices", "Premium OCR models", "All export formats", "24/7 support", "Auto-recharge", "API access", "Custom integrations"]
    )
}

class UserSubscription(BaseModel):
    user_id: str
    subscription_tier: str = "free"
    subscription_start_date: datetime
    subscription_end_date: Optional[datetime] = None
    credits: int = 100
    monthly_invoice_count: int = 0
    monthly_reset_date: datetime
    api_calls_this_hour: int = 0
    api_hour_reset: datetime
    auto_recharge_enabled: bool = False
    auto_recharge_threshold: int = 100
    auto_recharge_amount: int = 500
    is_active: bool = True

class TransactionStatus(str):
    PENDING = "pending"
    SUCCESS = "success"
    FAILED = "failed"
    REFUNDED = "refunded"

class Transaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    transaction_type: str  # "credit_purchase", "subscription_upgrade", "subscription_renewal"
    amount: float
    currency: str = "INR"
    credits_added: Optional[int] = None
    plan_upgraded_to: Optional[str] = None
    payment_method: str  # "razorpay", "stripe", "manual"
    payment_gateway_id: Optional[str] = None  # External transaction ID
    razorpay_order_id: Optional[str] = None
    razorpay_payment_id: Optional[str] = None
    razorpay_signature: Optional[str] = None
    status: str = TransactionStatus.PENDING
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completed_at: Optional[datetime] = None
    notes: Optional[str] = None

class TransactionOut(Transaction):
    pass

class UsageTracking(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    invoice_id: Optional[str] = None
    action_type: str  # "invoice_process", "export", "api_call"
    model_used: Optional[str] = None  # "haiku", "sonnet", "claude-3-5-sonnet"
    credits_consumed: int = 0
    processing_time_ms: Optional[int] = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    metadata: Optional[Dict[str, Any]] = None

class UsageStats(BaseModel):
    total_invoices_processed: int
    invoices_this_month: int
    total_credits_used: int
    credits_remaining: int
    subscription_tier: str
    monthly_limit: Optional[int]
    api_calls_remaining: int

class NotificationPreferences(BaseModel):
    payment_reminders: bool = True
    invoice_alerts: bool = True
    due_date_alerts: bool = True
    low_credit_alerts: bool = True
    push_enabled: bool = True
    email_enabled: bool = True

class UserProfileUpdate(BaseModel):
    name: Optional[str] = None
    business_name: Optional[str] = None
    gstin: Optional[str] = None
    business_type: Optional[str] = None

class PasswordChange(BaseModel):
    old_password: str
    new_password: str

class CreditPurchaseRequest(BaseModel):
    credits: int = Field(ge=100, le=10000)  # Min 100, Max 10000 credits per transaction
    payment_method: str = "razorpay"

class CreditPackage(BaseModel):
    credits: int
    price: float
    bonus_credits: int
    total_credits: int
    savings_percent: int

CREDIT_PACKAGES = [
    CreditPackage(credits=100, price=100.0, bonus_credits=0, total_credits=100, savings_percent=0),
    CreditPackage(credits=500, price=500.0, bonus_credits=50, total_credits=550, savings_percent=10),
    CreditPackage(credits=1000, price=1000.0, bonus_credits=200, total_credits=1200, savings_percent=20),
    CreditPackage(credits=5000, price=5000.0, bonus_credits=1500, total_credits=6500, savings_percent=30),
]

class SubscriptionUpgradeRequest(BaseModel):
    target_plan: str  # "starter" or "pro"
    payment_method: str = "razorpay"

# ─── Payment Models ───────────────────────────────────────────────────────────

class PaymentOrderRequest(BaseModel):
    amount: float  # Amount in rupees
    currency: str = "INR"
    type: str  # "credit_purchase" or "subscription_upgrade"
    credits: Optional[int] = None
    plan: Optional[str] = None
    notes: Optional[Dict[str, Any]] = None

class PaymentOrderResponse(BaseModel):
    order_id: str
    amount: int  # Amount in paise
    currency: str
    key_id: str  # Razorpay key for frontend
    transaction_id: str  # Our internal transaction ID

class PaymentVerificationRequest(BaseModel):
    razorpay_order_id: str
    razorpay_payment_id: str
    razorpay_signature: str
    transaction_id: str  # Our internal transaction ID

class PaymentVerificationResponse(BaseModel):
    success: bool
    transaction_id: str
    credits_added: Optional[int] = None
    new_balance: Optional[int] = None
    plan_upgraded: Optional[str] = None
    message: str

# ─── JWT Utilities ────────────────────────────────────────────────────────────

def create_token(user_id: str, email: str) -> str:
    expire = datetime.now(timezone.utc) + timedelta(days=JWT_EXPIRE_DAYS)
    payload = {"sub": user_id, "email": email, "exp": expire}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except JWTError:
        raise HTTPException(
            status_code=401,
            detail="Invalid or expired token. Please login again.",
            headers={"WWW-Authenticate": "Bearer"}
        )

async def get_current_user(request: Request, credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    payload = decode_token(credentials.credentials)
    user = await db.users.find_one({"id": payload["sub"]})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    request.state.user_id = payload["sub"]
    return user


def _estimate_base64_size(b64_text: str) -> int:
    if not b64_text:
        return 0
    if "base64," in b64_text:
        b64_text = b64_text.split("base64,", 1)[1]
    padding = b64_text.count("=")
    return max(int(len(b64_text) * 3 / 4) - padding, 0)

def hash_password(pwd: str) -> str:
    return bcrypt.hashpw(pwd.encode(), bcrypt.gensalt()).decode()

def verify_password(pwd: str, hashed: str) -> bool:
    return bcrypt.checkpw(pwd.encode(), hashed.encode())

# ─── Subscription & Credits Utilities ────────────────────────────────────────

async def get_or_create_subscription(user_id: str) -> dict:
    """Get user subscription or create default one"""
    subscription = await db.subscriptions.find_one({"user_id": user_id})
    if not subscription:
        now = datetime.now(timezone.utc)
        subscription = {
            "user_id": user_id,
            "subscription_tier": "free",
            "subscription_start_date": now,
            "subscription_end_date": None,
            "credits": PLAN_CONFIGS["free"].initial_credits,
            "monthly_invoice_count": 0,
            "monthly_reset_date": now.replace(day=1, hour=0, minute=0, second=0, microsecond=0) + timedelta(days=32),
            "api_calls_this_hour": 0,
            "api_hour_reset": now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1),
            "auto_recharge_enabled": False,
            "auto_recharge_threshold": 100,
            "auto_recharge_amount": 500,
            "is_active": True,
            "created_at": now,
            "updated_at": now
        }
        # Get next month's first day properly
        next_month = subscription["monthly_reset_date"].replace(day=1)
        if next_month.month == 12:
            next_month = next_month.replace(year=next_month.year + 1, month=1)
        else:
            next_month = next_month.replace(month=next_month.month + 1)
        subscription["monthly_reset_date"] = next_month
        await db.subscriptions.insert_one(subscription)
    return subscription

async def check_and_reset_monthly_limits(subscription: dict) -> dict:
    """Reset monthly counters if needed"""
    now = datetime.now(timezone.utc)
    reset_date = subscription["monthly_reset_date"]
    if reset_date.tzinfo is None:
        reset_date = reset_date.replace(tzinfo=timezone.utc)
    if now >= reset_date:
        next_month = subscription["monthly_reset_date"]
        if next_month.month == 12:
            next_month = next_month.replace(year=next_month.year + 1, month=1)
        else:
            next_month = next_month.replace(month=next_month.month + 1)
        
        await db.subscriptions.update_one(
            {"user_id": subscription["user_id"]},
            {
                "$set": {
                    "monthly_invoice_count": 0,
                    "monthly_reset_date": next_month,
                    "updated_at": now
                }
            }
        )
        subscription["monthly_invoice_count"] = 0
        subscription["monthly_reset_date"] = next_month
    return subscription

async def check_and_reset_hourly_limits(subscription: dict) -> dict:
    """Reset hourly API call counter if needed"""
    now = datetime.now(timezone.utc)
    hour_reset = subscription["api_hour_reset"]
    if hour_reset.tzinfo is None:
        hour_reset = hour_reset.replace(tzinfo=timezone.utc)
    if now >= hour_reset:
        next_hour = now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
        await db.subscriptions.update_one(
            {"user_id": subscription["user_id"]},
            {
                "$set": {
                    "api_calls_this_hour": 0,
                    "api_hour_reset": next_hour,
                    "updated_at": now
                }
            }
        )
        subscription["api_calls_this_hour"] = 0
        subscription["api_hour_reset"] = next_hour
    return subscription

async def check_subscription_limits(user_id: str, credits_needed: int = 0) -> tuple[bool, str, dict]:
    """
    Check if user can perform action based on subscription limits.
    Returns: (can_proceed, error_message, subscription)
    """
    subscription = await get_or_create_subscription(user_id)
    subscription = await check_and_reset_monthly_limits(subscription)
    subscription = await check_and_reset_hourly_limits(subscription)
    
    plan_config = PLAN_CONFIGS.get(subscription["subscription_tier"], PLAN_CONFIGS["free"])
    
    # Check if subscription is active (for paid plans)
    if subscription["subscription_tier"] != "free":
        if subscription.get("subscription_end_date"):
            if datetime.now(timezone.utc) > subscription["subscription_end_date"]:
                return False, "Subscription expired. Please renew your plan.", subscription
    
    # Check monthly invoice limit
    if plan_config.monthly_invoice_limit is not None:
        if subscription["monthly_invoice_count"] >= plan_config.monthly_invoice_limit:
            return False, f"Monthly invoice limit ({plan_config.monthly_invoice_limit}) reached. Upgrade your plan.", subscription
    
    # Check hourly API rate limit
    if subscription["api_calls_this_hour"] >= plan_config.api_calls_per_hour:
        return False, f"Hourly API limit ({plan_config.api_calls_per_hour}) reached. Please try again later.", subscription
    
    # Check credits
    if credits_needed > 0:
        if subscription["credits"] < credits_needed:
            return False, f"Insufficient credits. You need {credits_needed} but have {subscription['credits']}. Please purchase more credits.", subscription
    
    return True, "", subscription

async def deduct_credits(user_id: str, credits: int, invoice_id: Optional[str] = None, model_used: Optional[str] = None, action_type: str = "invoice_process") -> bool:
    """
    Deduct credits from user account and log usage.
    Returns True if successful, False if insufficient credits.
    """
    subscription = await get_or_create_subscription(user_id)
    
    if subscription["credits"] < credits:
        return False
    
    now = datetime.now(timezone.utc)
    
    # Deduct credits
    result = await db.subscriptions.update_one(
        {"user_id": user_id, "credits": {"$gte": credits}},  # Atomic check
        {
            "$inc": {"credits": -credits},
            "$set": {"updated_at": now}
        }
    )
    
    if result.modified_count == 0:
        return False
    
    # Log usage
    usage_record = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "invoice_id": invoice_id,
        "action_type": action_type,
        "model_used": model_used,
        "credits_consumed": credits,
        "timestamp": now
    }
    await db.usage_tracking.insert_one(usage_record)
    
    # Check auto-recharge
    if subscription.get("auto_recharge_enabled"):
        updated_sub = await db.subscriptions.find_one({"user_id": user_id})
        if updated_sub["credits"] <= subscription.get("auto_recharge_threshold", 100):
            # Trigger auto-recharge (would integrate with payment gateway)
            logger.info(f"Auto-recharge triggered for user {user_id}")
            # TODO: Implement actual auto-recharge logic
    
    return True

async def increment_api_call(user_id: str):
    """Increment hourly API call counter"""
    await db.subscriptions.update_one(
        {"user_id": user_id},
        {"$inc": {"api_calls_this_hour": 1}}
    )

async def increment_monthly_invoice_count(user_id: str):
    """Increment monthly invoice counter"""
    await db.subscriptions.update_one(
        {"user_id": user_id},
        {"$inc": {"monthly_invoice_count": 1}}
    )

def validate_gstin(gstin: Optional[str]) -> bool:
    """Validate Indian GSTIN format"""
    if not gstin:
        return True  # Optional field
    # GSTIN format: 2 digits (state) + 10 chars (PAN) + 1 digit (entity) + 1 char (Z) + 1 check digit
    pattern = r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$'
    return bool(re.match(pattern, gstin.upper()))

def validate_plan_tier(tier: str) -> bool:
    """Validate plan tier"""
    return tier in PLAN_CONFIGS

# ─── AI Extraction ────────────────────────────────────────────────────────────

EXTRACTION_SYSTEM_PROMPT = """You are an expert invoice data extraction AI specializing in Indian business invoices.
Extract all invoice fields from the provided image and return ONLY valid JSON — no markdown, no explanation.

Return this exact JSON structure:
{
  "invoice_number": "string or null",
  "invoice_date": "YYYY-MM-DD or null",
  "due_date": "YYYY-MM-DD or null",
  "vendor_name": "string or null",
  "vendor_gstin": "15-char string or null",
  "vendor_phone": "string or null",
  "vendor_address": "string or null",
  "buyer_name": "string or null",
  "buyer_gstin": "string or null",
  "line_items": [
    {
      "description": "string",
      "hsn_code": "string or null",
      "quantity": "number or null",
      "unit": "string or null",
      "unit_price": "number or null",
      "line_total": "number",
      "gst_rate": "number or null"
    }
  ],
  "subtotal": "number or null",
  "discount": "number or null",
  "cgst": "number or null",
  "sgst": "number or null",
  "igst": "number or null",
  "total_tax": "number or null",
  "grand_total": "number",
  "currency": "INR",
  "payment_terms": "string or null",
  "notes": "string or null",
  "confidence_scores": {
    "invoice_number": 0-100,
    "invoice_date": 0-100,
    "vendor_name": 0-100,
    "grand_total": 0-100,
    "line_items": 0-100
  }
}

Rules:
1. For intra-state GST: extract CGST and SGST. For inter-state: extract IGST only.
2. All monetary amounts should be numbers (not strings).
3. Dates must be YYYY-MM-DD format.
4. If text is Hindi/Gujarati/Tamil, extract data and transliterate vendor/buyer names to English.
5. Confidence 95-100=very clear, 70-94=readable, 40-69=partially visible, 0-39=unclear.
6. If image is not an invoice, return {"error": "Not an invoice"}.
7. Return null for missing fields, not empty strings."""

async def extract_invoice_from_image(image_base64: str, mime_type: str = "image/jpeg") -> dict:
    """Run AI extraction pipeline on invoice image."""
    try:
        llm_manager = _get_llm_manager()
        if not llm_manager.has_providers():
            raise ValueError("No LLM providers configured")

        provider, model = llm_manager.resolve_provider_and_model("premium")
        message = LLMMessage(
            text="Extract all invoice data from this image and return the JSON as specified.",
            images=[image_base64],
            image_mime_type=mime_type,
        )
        response = await llm_manager.complete(
            system_prompt=EXTRACTION_SYSTEM_PROMPT,
            user_message=message,
            model=model,
            provider=provider,
            enable_fallback=True,
        )

        # Strip markdown fences if present
        text = response.content.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
            text = text.strip()

        extracted = json.loads(text)
        if "error" in extracted:
            return {"error": extracted["error"]}
        return extracted

    except json.JSONDecodeError as e:
        logger.error(f"JSON parse error: {e}")
        return {"error": "Failed to parse extraction response"}
    except Exception as e:
        logger.error(f"Extraction error: {e}")
        return {"error": str(e)}

def validate_invoice_math(data: dict) -> list:
    """Validate invoice math and return list of issues."""
    issues = []
    try:
        subtotal = float(data.get("subtotal") or 0)
        discount = float(data.get("discount") or 0)
        cgst = float(data.get("cgst") or 0)
        sgst = float(data.get("sgst") or 0)
        igst = float(data.get("igst") or 0)
        total_tax = float(data.get("total_tax") or 0)
        grand_total = float(data.get("grand_total") or 0)

        # Check total_tax = cgst + sgst + igst
        computed_tax = cgst + sgst + igst
        if abs(computed_tax - total_tax) > 1.0 and total_tax > 0:
            issues.append({
                "field": "total_tax",
                "issue_type": "math_error",
                "message": f"Total tax ({total_tax}) ≠ CGST+SGST+IGST ({computed_tax:.2f})",
                "expected": str(computed_tax),
                "found": str(total_tax)
            })

        # Check grand_total = subtotal - discount + total_tax
        if subtotal > 0:
            computed_total = subtotal - discount + computed_tax
            if abs(computed_total - grand_total) > 2.0:
                issues.append({
                    "field": "grand_total",
                    "issue_type": "math_error",
                    "message": f"Grand total ({grand_total}) ≠ Subtotal-Discount+Tax ({computed_total:.2f})",
                    "expected": str(computed_total),
                    "found": str(grand_total)
                })

        # Validate line items sum
        line_items = data.get("line_items", [])
        if line_items and subtotal > 0:
            items_total = sum(float(item.get("line_total") or 0) for item in line_items)
            if abs(items_total - subtotal) > 2.0:
                issues.append({
                    "field": "subtotal",
                    "issue_type": "math_error",
                    "message": f"Line items total ({items_total:.2f}) ≠ subtotal ({subtotal})",
                    "expected": str(items_total),
                    "found": str(subtotal)
                })
    except Exception:
        pass
    return issues

async def detect_duplicates(user_id: str, data: dict) -> list:
    """Check for potential duplicate invoices."""
    candidates = []
    try:
        vendor = data.get("vendor_name", "")
        invoice_no = data.get("invoice_number", "")
        grand_total = float(data.get("grand_total") or 0)
        invoice_date = data.get("invoice_date", "")

        query = {"user_id": user_id}
        if vendor:
            safe_vendor = re.escape(vendor)
            query["vendor_name"] = {"$regex": f"^{safe_vendor}$", "$options": "i"}

        existing = await db.invoices.find(query).limit(50).to_list(50)
        for inv in existing:
            reasons = []
            score = 0.0
            if invoice_no and inv.get("invoice_number") == invoice_no:
                reasons.append("same_invoice_number")
                score += 0.5
            if grand_total > 0 and abs(float(inv.get("grand_total", 0)) - grand_total) < 1.0:
                reasons.append("same_amount")
                score += 0.3
            if invoice_date and inv.get("invoice_date") == invoice_date:
                reasons.append("same_date")
                score += 0.2
            if score >= 0.5:
                candidates.append({
                    "invoice_id": inv["id"],
                    "similarity_score": round(score, 2),
                    "match_reasons": reasons,
                    "vendor_name": inv.get("vendor_name"),
                    "invoice_number": inv.get("invoice_number"),
                    "grand_total": inv.get("grand_total"),
                    "invoice_date": inv.get("invoice_date")
                })
    except Exception as e:
        logger.error(f"Duplicate detection error: {e}")
    return candidates

# ─── Auth Routes ──────────────────────────────────────────────────────────────

@api_router.post("/auth/register")
@limiter.limit(RATE_LIMIT_AUTH)
async def register(request: Request, data: UserCreate):
    existing = await db.users.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Validate GSTIN if provided
    if data.gstin and not validate_gstin(data.gstin):
        raise HTTPException(status_code=400, detail="Invalid GSTIN format")
    
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    user_doc = {
        "id": user_id,
        "email": data.email.lower(),
        "password_hash": hash_password(data.password),
        "name": data.name,
        "business_name": data.business_name,
        "gstin": data.gstin.upper() if data.gstin else None,
        "business_type": data.business_type,
        "plan": "free",
        "notification_preferences": NotificationPreferences().dict(),
        "created_at": now,
        "updated_at": now
    }
    await db.users.insert_one(user_doc)
    
    # Create default subscription
    await get_or_create_subscription(user_id)
    
    token = create_token(user_id, data.email)
    return {
        "token": token,
        "user": {
            "id": user_id,
            "email": data.email,
            "name": data.name,
            "business_name": data.business_name,
            "plan": "free",
            "credits": PLAN_CONFIGS["free"].initial_credits,
            "monthly_invoice_limit": PLAN_CONFIGS["free"].monthly_invoice_limit
        }
    }

@api_router.post("/auth/login")
@limiter.limit(RATE_LIMIT_AUTH)
async def login(request: Request, data: UserLogin):
    user = await db.users.find_one({"email": data.email.lower()})
    if not user or not verify_password(data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Get subscription info
    subscription = await get_or_create_subscription(user["id"])
    plan_config = PLAN_CONFIGS.get(subscription["subscription_tier"], PLAN_CONFIGS["free"])
    
    token = create_token(user["id"], user["email"])
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "name": user["name"],
            "business_name": user["business_name"],
            "plan": user.get("plan", "free"),
            "credits": subscription["credits"],
            "monthly_invoice_count": subscription["monthly_invoice_count"],
            "monthly_invoice_limit": plan_config.monthly_invoice_limit
        }
    }

@api_router.get("/auth/me")
async def me(current_user: dict = Depends(get_current_user)):
    # Get subscription info
    subscription = await get_or_create_subscription(current_user["id"])
    subscription = await check_and_reset_monthly_limits(subscription)
    plan_config = PLAN_CONFIGS.get(subscription["subscription_tier"], PLAN_CONFIGS["free"])
    
    return {
        "id": current_user["id"],
        "email": current_user["email"],
        "name": current_user["name"],
        "business_name": current_user["business_name"],
        "gstin": current_user.get("gstin"),
        "business_type": current_user.get("business_type"),
        "plan": current_user.get("plan", "free"),
        "credits": subscription["credits"],
        "monthly_invoice_count": subscription["monthly_invoice_count"],
        "monthly_invoice_limit": plan_config.monthly_invoice_limit,
        "created_at": current_user["created_at"]
    }

# ─── Invoice Processing ───────────────────────────────────────────────────────

@api_router.post("/process-invoice")
@limiter.limit(RATE_LIMIT_PROCESS)
async def process_invoice(request: Request, req: ProcessInvoiceRequest, current_user: dict = Depends(get_current_user)):
    if req.mime_type not in ALLOWED_MIME_TYPES:
        raise HTTPException(status_code=415, detail="Unsupported file type")

    if _estimate_base64_size(req.image_base64) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Upload exceeds 10MB limit")

    if not _get_llm_manager().has_providers():
        raise HTTPException(status_code=500, detail="No LLM providers configured")

    # Check subscription limits BEFORE processing
    estimated_credits = 15  # Estimate, will adjust after knowing actual model used
    can_proceed, error_msg, subscription = await check_subscription_limits(current_user["id"], estimated_credits)
    
    if not can_proceed:
        raise HTTPException(status_code=403, detail=error_msg)
    
    # Increment API call counter
    await increment_api_call(current_user["id"])

    # Run extraction
    extracted = await extract_invoice_from_image(req.image_base64, req.mime_type)
    if "error" in extracted:
        raise HTTPException(status_code=422, detail=extracted["error"])

    # Determine credits needed (based on complexity - can be improved with actual model tracking)
    credits_needed = 10  # Default credit cost for invoice processing

    # Validate math
    validation_issues = validate_invoice_math(extracted)

    # Detect duplicates
    duplicate_candidates = await detect_duplicates(current_user["id"], extracted)

    # Extract confidence scores
    confidence_scores = extracted.pop("confidence_scores", {})
    overall_confidence = int(sum(confidence_scores.values()) / max(len(confidence_scores), 1)) if confidence_scores else 80

    # Deduct credits after successful processing
    credit_deducted = await deduct_credits(
        current_user["id"],
        credits_needed,
        invoice_id=None,  # Will be set when invoice is saved
        model_used="claude-vision",
        action_type="invoice_process"
    )
    
    if not credit_deducted:
        logger.warning(f"Credit deduction failed for user {current_user['id']}")
    
    # Increment monthly invoice count
    await increment_monthly_invoice_count(current_user["id"])

    return {
        "extracted_data": extracted,
        "confidence_scores": confidence_scores,
        "overall_confidence": overall_confidence,
        "validation_issues": validation_issues,
        "duplicate_candidates": duplicate_candidates,
        "source_type": req.source_type,
        "credits_used": credits_needed,
        "credits_remaining": subscription["credits"] - credits_needed
    }

# ─── Invoice CRUD ─────────────────────────────────────────────────────────────

@api_router.post("/invoices", response_model=InvoiceOut)
async def create_invoice(data: InvoiceCreate, current_user: dict = Depends(get_current_user)):
    invoice_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)

    # Upsert vendor
    if data.vendor_name:
        await db.vendors.update_one(
            {"user_id": current_user["id"], "name": {"$regex": f"^{data.vendor_name}$", "$options": "i"}},
            {"$setOnInsert": {
                "id": str(uuid.uuid4()),
                "user_id": current_user["id"],
                "name": data.vendor_name,
                "gstin": data.vendor_gstin,
                "phone": data.vendor_phone,
                "address": data.vendor_address,
                "created_at": now
            }, "$inc": {"total_invoices": 1, "total_spend": float(data.grand_total or 0)}},
            upsert=True
        )

    invoice_doc = {
        "id": invoice_id,
        "user_id": current_user["id"],
        **data.dict(),
        "line_items": [item.dict() for item in data.line_items],
        "created_at": now,
        "updated_at": now
    }
    await db.invoices.insert_one(invoice_doc)
    return {**invoice_doc}

@api_router.get("/invoices")
async def list_invoices(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    status_filter: Optional[str] = None,
    vendor_name: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["id"]}
    if search:
        safe_search = re.escape(search)
        query["$or"] = [
            {"vendor_name": {"$regex": safe_search, "$options": "i"}},
            {"invoice_number": {"$regex": safe_search, "$options": "i"}},
            {"buyer_name": {"$regex": safe_search, "$options": "i"}}
        ]
    if status_filter:
        query["status"] = status_filter
    if vendor_name:
        query["vendor_name"] = {"$regex": re.escape(vendor_name), "$options": "i"}
    if date_from:
        query["invoice_date"] = query.get("invoice_date", {})
        query["invoice_date"]["$gte"] = date_from
    if date_to:
        query["invoice_date"] = query.get("invoice_date", {})
        query["invoice_date"]["$lte"] = date_to

    total = await db.invoices.count_documents(query)
    invoices = await db.invoices.find(query).sort("created_at", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    for inv in invoices:
        inv.pop("_id", None)

    # Summary
    pipeline = [
        {"$match": {"user_id": current_user["id"]}},
        {"$group": {
            "_id": None,
            "total_amount": {"$sum": "$grand_total"},
            "pending_amount": {"$sum": {"$cond": [{"$eq": ["$status", "pending"]}, "$grand_total", 0]}},
            "overdue_amount": {"$sum": {"$cond": [{"$eq": ["$status", "overdue"]}, "$grand_total", 0]}}
        }}
    ]
    summary_result = await db.invoices.aggregate(pipeline).to_list(1)
    summary = summary_result[0] if summary_result else {"total_amount": 0, "pending_amount": 0, "overdue_amount": 0}
    summary.pop("_id", None)

    return {"invoices": invoices, "total": total, "page": page, "limit": limit, "summary": summary}

@api_router.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    inv = await db.invoices.find_one({"id": invoice_id, "user_id": current_user["id"]})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    inv.pop("_id", None)
    return inv

@api_router.put("/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, data: InvoiceCreate, current_user: dict = Depends(get_current_user)):
    inv = await db.invoices.find_one({"id": invoice_id, "user_id": current_user["id"]})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    update_data = {**data.dict(), "line_items": [item.dict() for item in data.line_items], "updated_at": datetime.now(timezone.utc)}
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    updated = await db.invoices.find_one({"id": invoice_id})
    updated.pop("_id", None)
    return updated

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    inv = await db.invoices.find_one({"id": invoice_id, "user_id": current_user["id"]})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    await db.invoices.delete_one({"id": invoice_id})
    return {"message": "Invoice deleted"}

@api_router.patch("/invoices/{invoice_id}/status")
async def update_invoice_status(invoice_id: str, data: InvoiceStatusUpdate, current_user: dict = Depends(get_current_user)):
    inv = await db.invoices.find_one({"id": invoice_id, "user_id": current_user["id"]})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    await db.invoices.update_one({"id": invoice_id}, {"$set": {"status": data.status, "updated_at": datetime.now(timezone.utc)}})
    return {"message": "Status updated", "status": data.status}

# ─── Vendors ─────────────────────────────────────────────────────────────────

@api_router.get("/vendors")
async def list_vendors(current_user: dict = Depends(get_current_user)):
    pipeline = [
        {"$match": {"user_id": current_user["id"]}},
        {"$group": {
            "_id": "$vendor_name",
            "total_invoices": {"$sum": 1},
            "total_spend": {"$sum": "$grand_total"},
            "last_invoice_date": {"$max": "$invoice_date"},
            "vendor_gstin": {"$first": "$vendor_gstin"},
            "vendor_phone": {"$first": "$vendor_phone"}
        }},
        {"$sort": {"total_spend": -1}}
    ]
    vendors = await db.invoices.aggregate(pipeline).to_list(100)
    return [
        {
            "name": v["_id"],
            "total_invoices": v["total_invoices"],
            "total_spend": round(v["total_spend"], 2),
            "last_invoice_date": v["last_invoice_date"],
            "gstin": v.get("vendor_gstin"),
            "phone": v.get("vendor_phone")
        }
        for v in vendors if v["_id"]
    ]

# ─── Analytics ────────────────────────────────────────────────────────────────

@api_router.get("/analytics/summary")
async def analytics_summary(current_user: dict = Depends(get_current_user)):
    user_id = current_user["id"]
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).strftime("%Y-%m-%d")

    # Overall stats
    all_pipeline = [
        {"$match": {"user_id": user_id}},
        {"$group": {
            "_id": None,
            "total_invoices": {"$sum": 1},
            "total_amount": {"$sum": "$grand_total"},
            "paid_amount": {"$sum": {"$cond": [{"$eq": ["$status", "paid"]}, "$grand_total", 0]}},
            "pending_amount": {"$sum": {"$cond": [{"$eq": ["$status", "pending"]}, "$grand_total", 0]}},
            "overdue_amount": {"$sum": {"$cond": [{"$eq": ["$status", "overdue"]}, "$grand_total", 0]}}
        }}
    ]
    all_result = await db.invoices.aggregate(all_pipeline).to_list(1)
    stats = all_result[0] if all_result else {"total_invoices": 0, "total_amount": 0, "paid_amount": 0, "pending_amount": 0, "overdue_amount": 0}
    stats.pop("_id", None)

    # This month stats
    month_pipeline = [
        {"$match": {"user_id": user_id, "invoice_date": {"$gte": month_start}}},
        {"$group": {"_id": None, "count": {"$sum": 1}, "amount": {"$sum": "$grand_total"}}}
    ]
    month_result = await db.invoices.aggregate(month_pipeline).to_list(1)
    month_stats = month_result[0] if month_result else {"count": 0, "amount": 0}
    month_stats.pop("_id", None)

    # Top vendors
    vendor_pipeline = [
        {"$match": {"user_id": user_id}},
        {"$group": {"_id": "$vendor_name", "total_spend": {"$sum": "$grand_total"}, "count": {"$sum": 1}}},
        {"$sort": {"total_spend": -1}},
        {"$limit": 5}
    ]
    top_vendors = await db.invoices.aggregate(vendor_pipeline).to_list(5)
    top_vendors = [{"name": v["_id"], "total_spend": round(v["total_spend"], 2), "count": v["count"]} for v in top_vendors if v["_id"]]

    # Monthly trend (last 6 months)
    monthly_trend = []
    for i in range(5, -1, -1):
        month_date = now.replace(day=1) - timedelta(days=i * 30)
        m_start = month_date.replace(day=1).strftime("%Y-%m-%d")
        if month_date.month == 12:
            m_end = month_date.replace(year=month_date.year + 1, month=1, day=1).strftime("%Y-%m-%d")
        else:
            m_end = month_date.replace(month=month_date.month + 1, day=1).strftime("%Y-%m-%d")
        trend_result = await db.invoices.aggregate([
            {"$match": {"user_id": user_id, "invoice_date": {"$gte": m_start, "$lt": m_end}}},
            {"$group": {"_id": None, "amount": {"$sum": "$grand_total"}, "count": {"$sum": 1}}}
        ]).to_list(1)
        trend_data = trend_result[0] if trend_result else {"amount": 0, "count": 0}
        monthly_trend.append({"month": month_date.strftime("%b %Y"), "amount": round(trend_data.get("amount", 0), 2), "count": trend_data.get("count", 0)})

    # Category breakdown
    category_pipeline = [
        {"$match": {"user_id": user_id, "category": {"$ne": None}}},
        {"$group": {"_id": "$category", "amount": {"$sum": "$grand_total"}, "count": {"$sum": 1}}},
        {"$sort": {"amount": -1}}
    ]
    categories = await db.invoices.aggregate(category_pipeline).to_list(10)
    category_data = [{"category": c["_id"], "amount": round(c["amount"], 2), "count": c["count"]} for c in categories]

    # Status breakdown
    status_pipeline = [
        {"$match": {"user_id": user_id}},
        {"$group": {"_id": "$status", "count": {"$sum": 1}, "amount": {"$sum": "$grand_total"}}}
    ]
    status_data = await db.invoices.aggregate(status_pipeline).to_list(10)
    status_breakdown = {s["_id"]: {"count": s["count"], "amount": round(s["amount"], 2)} for s in status_data if s["_id"]}

    return {
        "overall": stats,
        "this_month": month_stats,
        "top_vendors": top_vendors,
        "monthly_trend": monthly_trend,
        "categories": category_data,
        "status_breakdown": status_breakdown
    }

# ─── Export ───────────────────────────────────────────────────────────────────

@api_router.post("/export")
@limiter.limit(RATE_LIMIT_EXPORT)
async def export_invoices(request: Request, req: ExportRequest, current_user: dict = Depends(get_current_user)):
    query = {"user_id": current_user["id"]}
    if req.date_from:
        query["invoice_date"] = query.get("invoice_date", {})
        query["invoice_date"]["$gte"] = req.date_from
    if req.date_to:
        query["invoice_date"] = query.get("invoice_date", {})
        query["invoice_date"]["$lte"] = req.date_to
    if req.vendor_name:
        query["vendor_name"] = {"$regex": req.vendor_name, "$options": "i"}
    if req.status_filter:
        query["status"] = req.status_filter

    total = await db.invoices.count_documents(query)
    if total > EXPORT_MAX_RECORDS:
        raise HTTPException(
            status_code=413,
            detail=f"Export exceeds limit of {EXPORT_MAX_RECORDS} invoices"
        )

    invoices = await db.invoices.find(query).sort("invoice_date", -1).to_list(EXPORT_MAX_RECORDS)

    if req.format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        headers = ["Invoice ID", "Invoice No", "Date", "Due Date", "Vendor", "Vendor GSTIN",
                   "Subtotal", "Discount", "CGST", "SGST", "IGST", "Total Tax", "Grand Total",
                   "Currency", "Status", "Category", "Source", "Notes"]
        writer.writerow(headers)
        for inv in invoices:
            writer.writerow([
                inv.get("id"), inv.get("invoice_number"), inv.get("invoice_date"),
                inv.get("due_date"), inv.get("vendor_name"), inv.get("vendor_gstin"),
                inv.get("subtotal"), inv.get("discount"), inv.get("cgst"), inv.get("sgst"),
                inv.get("igst"), inv.get("total_tax"), inv.get("grand_total"),
                inv.get("currency", "INR"), inv.get("status"), inv.get("category"),
                inv.get("source_type"), inv.get("notes")
            ])
        csv_bytes = output.getvalue().encode("utf-8")
        b64 = base64.b64encode(csv_bytes).decode()
        return {"format": "csv", "filename": f"InvoiceAI_Export.csv", "data_base64": b64, "count": len(invoices)}

    else:  # Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1E1B4B")
        alt_fill = PatternFill("solid", fgColor="F3F4F6")
        center = Alignment(horizontal="center")

        headers = ["Invoice ID", "Invoice No", "Date", "Due Date", "Vendor Name", "Vendor GSTIN",
                   "Buyer Name", "Subtotal (₹)", "Discount (₹)", "CGST (₹)", "SGST (₹)",
                   "IGST (₹)", "Total Tax (₹)", "Grand Total (₹)", "Currency", "Status",
                   "Category", "Source", "Notes", "Created At"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for row_idx, inv in enumerate(invoices, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            row_data = [
                inv.get("id"), inv.get("invoice_number"), inv.get("invoice_date"),
                inv.get("due_date"), inv.get("vendor_name"), inv.get("vendor_gstin"),
                inv.get("buyer_name"), inv.get("subtotal"), inv.get("discount"),
                inv.get("cgst"), inv.get("sgst"), inv.get("igst"), inv.get("total_tax"),
                inv.get("grand_total"), inv.get("currency", "INR"), inv.get("status"),
                inv.get("category"), inv.get("source_type"), inv.get("notes"),
                inv.get("created_at").strftime("%Y-%m-%d") if isinstance(inv.get("created_at"), datetime) else str(inv.get("created_at", ""))
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                if fill:
                    cell.fill = fill

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        # Line items sheet
        ws2 = wb.create_sheet("Line Items")
        li_headers = ["Invoice ID", "Description", "HSN Code", "Qty", "Unit", "Unit Price", "Line Total", "GST %"]
        for col, h in enumerate(li_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        li_row = 2
        for inv in invoices:
            for item in inv.get("line_items", []):
                ws2.cell(row=li_row, column=1, value=inv.get("id"))
                ws2.cell(row=li_row, column=2, value=item.get("description"))
                ws2.cell(row=li_row, column=3, value=item.get("hsn_code"))
                ws2.cell(row=li_row, column=4, value=item.get("quantity"))
                ws2.cell(row=li_row, column=5, value=item.get("unit"))
                ws2.cell(row=li_row, column=6, value=item.get("unit_price"))
                ws2.cell(row=li_row, column=7, value=item.get("line_total"))
                ws2.cell(row=li_row, column=8, value=item.get("gst_rate"))
                li_row += 1

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        b64 = base64.b64encode(buffer.read()).decode()
        return {"format": "excel", "filename": "InvoiceAI_Export.xlsx", "data_base64": b64, "count": len(invoices)}

# ─── Corrections ─────────────────────────────────────────────────────────────

@api_router.post("/corrections")
async def log_correction(data: CorrectionCreate, current_user: dict = Depends(get_current_user)):
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        **data.dict(),
        "created_at": datetime.now(timezone.utc)
    }
    await db.corrections.insert_one(doc)
    return {"message": "Correction logged"}

# ─── Subscription & Credits APIs ──────────────────────────────────────────────

@api_router.get("/plans")
async def get_plans():
    """Get all available subscription plans"""
    return {
        "plans": [
            {
                "tier": plan.tier,
                "name": plan.name,
                "price_monthly": plan.price_monthly,
                "currency": plan.currency,
                "monthly_invoice_limit": plan.monthly_invoice_limit,
                "api_calls_per_hour": plan.api_calls_per_hour,
                "initial_credits": plan.initial_credits,
                "features": plan.features
            }
            for plan in PLAN_CONFIGS.values()
        ],
        "credit_packages": [pkg.dict() for pkg in CREDIT_PACKAGES]
    }

@api_router.get("/user/subscription")
@limiter.limit(RATE_LIMIT_DEFAULT)
async def get_user_subscription(request: Request, current_user: dict = Depends(get_current_user)):
    """Get current user subscription details"""
    subscription = await get_or_create_subscription(current_user["id"])
    subscription = await check_and_reset_monthly_limits(subscription)
    subscription = await check_and_reset_hourly_limits(subscription)
    
    plan_config = PLAN_CONFIGS.get(subscription["subscription_tier"], PLAN_CONFIGS["free"])
    
    return {
        "subscription_tier": subscription["subscription_tier"],
        "plan_name": plan_config.name,
        "credits": subscription["credits"],
        "monthly_invoice_count": subscription["monthly_invoice_count"],
        "monthly_invoice_limit": plan_config.monthly_invoice_limit,
        "api_calls_per_hour": plan_config.api_calls_per_hour,
        "api_calls_this_hour": subscription["api_calls_this_hour"],
        "subscription_start_date": subscription["subscription_start_date"],
        "subscription_end_date": subscription.get("subscription_end_date"),
        "monthly_reset_date": subscription["monthly_reset_date"],
        "auto_recharge_enabled": subscription.get("auto_recharge_enabled", False),
        "is_active": subscription.get("is_active", True),
        "features": plan_config.features
    }

@api_router.get("/user/usage-stats")
@limiter.limit(RATE_LIMIT_DEFAULT)
async def get_usage_stats(request: Request, current_user: dict = Depends(get_current_user)):
    """Get user usage statistics"""
    subscription = await get_or_create_subscription(current_user["id"])
    subscription = await check_and_reset_monthly_limits(subscription)
    
    plan_config = PLAN_CONFIGS.get(subscription["subscription_tier"], PLAN_CONFIGS["free"])
    
    # Get total invoices processed
    total_invoices = await db.invoices.count_documents({"user_id": current_user["id"]})
    
    # Get total credits used
    pipeline = [
        {"$match": {"user_id": current_user["id"]}},
        {"$group": {"_id": None, "total": {"$sum": "$credits_consumed"}}}
    ]
    credits_used = await db.usage_tracking.aggregate(pipeline).to_list(1)
    total_credits_used = credits_used[0]["total"] if credits_used else 0
    
    return {
        "total_invoices_processed": total_invoices,
        "invoices_this_month": subscription["monthly_invoice_count"],
        "total_credits_used": total_credits_used,
        "credits_remaining": subscription["credits"],
        "subscription_tier": subscription["subscription_tier"],
        "monthly_limit": plan_config.monthly_invoice_limit,
        "api_calls_remaining": max(0, plan_config.api_calls_per_hour - subscription["api_calls_this_hour"])
    }

@api_router.patch("/user/profile")
@limiter.limit(RATE_LIMIT_DEFAULT)
async def update_profile(
    request: Request,
    profile_update: UserProfileUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update user profile"""
    update_data = {}
    
    if profile_update.name:
        update_data["name"] = profile_update.name
    
    if profile_update.business_name:
        update_data["business_name"] = profile_update.business_name
    
    if profile_update.gstin is not None:
        if profile_update.gstin and not validate_gstin(profile_update.gstin):
            raise HTTPException(status_code=400, detail="Invalid GSTIN format")
        update_data["gstin"] = profile_update.gstin.upper() if profile_update.gstin else None
    
    if profile_update.business_type:
        update_data["business_type"] = profile_update.business_type
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": update_data}
    )
    
    return {"message": "Profile updated successfully", "updated_fields": list(update_data.keys())}

@api_router.patch("/user/password")
@limiter.limit("3/hour")  # Stricter limit for password changes
async def change_password(
    request: Request,
    password_change: PasswordChange,
    current_user: dict = Depends(get_current_user)
):
    """Change user password"""
    # Verify old password
    if not verify_password(password_change.old_password, current_user["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Validate new password
    if len(password_change.new_password) < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters")
    
    if password_change.new_password == password_change.old_password:
        raise HTTPException(status_code=400, detail="New password must be different from old password")
    
    # Update password
    new_hash = hash_password(password_change.new_password)
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"password_hash": new_hash, "updated_at": datetime.now(timezone.utc)}}
    )
    
    logger.info(f"Password changed for user {current_user['id']}")
    return {"message": "Password changed successfully"}

@api_router.patch("/user/notifications")
@limiter.limit(RATE_LIMIT_DEFAULT)
async def update_notifications(
    request: Request,
    preferences: NotificationPreferences,
    current_user: dict = Depends(get_current_user)
):
    """Update notification preferences"""
    await db.users.update_one(
        {"id": current_user["id"]},
        {
            "$set": {
                "notification_preferences": preferences.dict(),
                "updated_at": datetime.now(timezone.utc)
            }
        }
    )
    return {"message": "Notification preferences updated"}

@api_router.post("/credits/purchase")
@limiter.limit("10/hour")  # Prevent abuse
async def purchase_credits(
    request: Request,
    purchase_request: CreditPurchaseRequest,
    current_user: dict = Depends(get_current_user)
):
    """Initiate credit purchase"""
    # Find matching package or calculate custom amount
    matching_package = None
    for pkg in CREDIT_PACKAGES:
        if pkg.credits == purchase_request.credits:
            matching_package = pkg
            break
    
    if not matching_package:
        # Custom amount - no bonus
        amount = float(purchase_request.credits)
        total_credits = purchase_request.credits
    else:
        amount = matching_package.price
        total_credits = matching_package.total_credits
    
    # Create transaction record
    transaction = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "transaction_type": "credit_purchase",
        "amount": amount,
        "currency": "INR",
        "credits_added": total_credits,
        "payment_method": purchase_request.payment_method,
        "status": TransactionStatus.PENDING,
        "created_at": datetime.now(timezone.utc)
    }
    
    # For Razorpay integration (placeholder)
    if purchase_request.payment_method == "razorpay":
        # In production, create Razorpay order here
        razorpay_order_id = f"order_{uuid.uuid4().hex[:16]}"
        transaction["razorpay_order_id"] = razorpay_order_id
        
        await db.transactions.insert_one(transaction)
        
        return {
            "transaction_id": transaction["id"],
            "amount": amount,
            "currency": "INR",
            "credits": total_credits,
            "razorpay_order_id": razorpay_order_id,
            "message": "Complete payment with Razorpay"
        }
    
    await db.transactions.insert_one(transaction)
    return {
        "transaction_id": transaction["id"],
        "amount": amount,
        "credits": total_credits,
        "message": "Transaction created"
    }

@api_router.post("/credits/webhook")
async def credits_webhook(request: Request):
    """Handle payment gateway webhooks (Razorpay/Stripe)"""
    # This would verify webhook signature and process payment
    # For now, it's a placeholder for payment integration
    body = await request.json()
    
    # Verify signature (Razorpay example)
    # signature = request.headers.get("X-Razorpay-Signature")
    # if not verify_razorpay_signature(body, signature):
    #     raise HTTPException(status_code=400, detail="Invalid signature")
    
    # Process successful payment
    if body.get("event") == "payment.captured":
        payment_id = body.get("payload", {}).get("payment", {}).get("entity", {}).get("id")
        order_id = body.get("payload", {}).get("payment", {}).get("entity", {}).get("order_id")
        
        # Find transaction
        transaction = await db.transactions.find_one({"razorpay_order_id": order_id})
        if transaction:
            # Update transaction status
            await db.transactions.update_one(
                {"id": transaction["id"]},
                {
                    "$set": {
                        "status": TransactionStatus.SUCCESS,
                        "razorpay_payment_id": payment_id,
                        "completed_at": datetime.now(timezone.utc)
                    }
                }
            )
            
            # Add credits to user account
            await db.subscriptions.update_one(
                {"user_id": transaction["user_id"]},
                {"$inc": {"credits": transaction["credits_added"]}}
            )
            
            logger.info(f"Credits added: {transaction['credits_added']} for user {transaction['user_id']}")
    
    return {"status": "ok"}

@api_router.get("/credits/balance")
@limiter.limit(RATE_LIMIT_DEFAULT)
async def get_credit_balance(request: Request, current_user: dict = Depends(get_current_user)):
    """Get current credit balance"""
    subscription = await get_or_create_subscription(current_user["id"])
    return {
        "credits": subscription["credits"],
        "auto_recharge_enabled": subscription.get("auto_recharge_enabled", False),
        "auto_recharge_threshold": subscription.get("auto_recharge_threshold", 100)
    }

@api_router.get("/transactions")
@limiter.limit(RATE_LIMIT_DEFAULT)
async def get_transactions(
    request: Request,
    current_user: dict = Depends(get_current_user),
    limit: int = Query(20, ge=1, le=100)
):
    """Get transaction history"""
    transactions = await db.transactions.find(
        {"user_id": current_user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return {"transactions": transactions}

@api_router.post("/subscription/upgrade")
@limiter.limit("5/hour")
async def upgrade_subscription(
    request: Request,
    upgrade_request: SubscriptionUpgradeRequest,
    current_user: dict = Depends(get_current_user)
):
    """Upgrade subscription plan"""
    if not validate_plan_tier(upgrade_request.target_plan):
        raise HTTPException(status_code=400, detail="Invalid plan tier")
    
    if upgrade_request.target_plan == "free":
        raise HTTPException(status_code=400, detail="Cannot upgrade to free plan")
    
    subscription = await get_or_create_subscription(current_user["id"])
    current_tier = subscription["subscription_tier"]
    
    # Check if already on target plan
    if current_tier == upgrade_request.target_plan:
        raise HTTPException(status_code=400, detail="Already on this plan")
    
    # Prevent downgrade via upgrade endpoint
    tier_order = {"free": 0, "starter": 1, "pro": 2}
    if tier_order.get(upgrade_request.target_plan, 0) < tier_order.get(current_tier, 0):
        raise HTTPException(status_code=400, detail="Use downgrade endpoint to downgrade plan")
    
    plan_config = PLAN_CONFIGS[upgrade_request.target_plan]
    
    # Create transaction
    transaction = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "transaction_type": "subscription_upgrade",
        "amount": plan_config.price_monthly,
        "currency": plan_config.currency,
        "plan_upgraded_to": upgrade_request.target_plan,
        "payment_method": upgrade_request.payment_method,
        "status": TransactionStatus.PENDING,
        "created_at": datetime.now(timezone.utc)
    }
    
    if upgrade_request.payment_method == "razorpay":
        razorpay_order_id = f"order_{uuid.uuid4().hex[:16]}"
        transaction["razorpay_order_id"] = razorpay_order_id
        
        await db.transactions.insert_one(transaction)
        
        return {
            "transaction_id": transaction["id"],
            "amount": plan_config.price_monthly,
            "currency": plan_config.currency,
            "plan": upgrade_request.target_plan,
            "razorpay_order_id": razorpay_order_id
        }
    
    await db.transactions.insert_one(transaction)
    return {
        "transaction_id": transaction["id"],
        "amount": plan_config.price_monthly,
        "plan": upgrade_request.target_plan
    }

@api_router.get("/usage/history")
@limiter.limit(RATE_LIMIT_DEFAULT)
async def get_usage_history(
    request: Request,
    current_user: dict = Depends(get_current_user),
    limit: int = Query(50, ge=1, le=200),
    days: int = Query(30, ge=1, le=90)
):
    """Get usage history"""
    cutoff_date = datetime.now(timezone.utc) - timedelta(days=days)
    
    usage_records = await db.usage_tracking.find(
        {
            "user_id": current_user["id"],
            "timestamp": {"$gte": cutoff_date}
        },
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit).to_list(limit)
    
    # Aggregate by action type
    pipeline = [
        {"$match": {"user_id": current_user["id"], "timestamp": {"$gte": cutoff_date}}},
        {
            "$group": {
                "_id": "$action_type",
                "total_credits": {"$sum": "$credits_consumed"},
                "count": {"$sum": 1}
            }
        }
    ]
    aggregated = await db.usage_tracking.aggregate(pipeline).to_list(10)
    
    return {
        "usage_records": usage_records,
        "summary": aggregated
    }

# ═══════════════════════════════════════════════════════════════════════════════
# PAYMENT ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.post("/payment/create-order", response_model=PaymentOrderResponse)
@limiter.limit(RATE_LIMIT_DEFAULT)
async def create_payment_order(
    request: Request,
    order_request: PaymentOrderRequest,
    current_user: dict = Depends(get_current_user)
):
    """Create a payment order for credit purchase or subscription upgrade"""
    
    # Validate payment gateway is configured
    if not is_razorpay_configured():
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Payment gateway not configured. Please contact support."
        )
    
    # Validate order type
    if order_request.type not in ["credit_purchase", "subscription_upgrade"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid order type. Must be 'credit_purchase' or 'subscription_upgrade'"
        )
    
    # Validate amount
    if order_request.amount < 100:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Minimum payment amount is ₹100"
        )
    
    # Create internal transaction record first
    transaction_id = str(uuid.uuid4())
    transaction = {
        "id": transaction_id,
        "user_id": current_user["id"],
        "amount": order_request.amount,
        "currency": order_request.currency,
        "payment_method": "razorpay",
        "status": TransactionStatus.PENDING,
        "created_at": datetime.now(timezone.utc),
        "notes": f"Order type: {order_request.type}"
    }
    
    # Add type-specific fields
    if order_request.type == "credit_purchase" and order_request.credits:
        transaction["credits_added"] = order_request.credits
    elif order_request.type == "subscription_upgrade" and order_request.plan:
        transaction["plan_upgraded_to"] = order_request.plan
    
    # Create Razorpay order
    try:
        amount_in_paise = rupees_to_paise(order_request.amount)
        razorpay_order = create_razorpay_order(
            amount=amount_in_paise,
            currency=order_request.currency,
            receipt=transaction_id,
            notes={
                "user_id": current_user["id"],
                "user_email": current_user["email"],
                "type": order_request.type,
                **(order_request.notes or {})
            }
        )
        
        # Update transaction with Razorpay order ID
        transaction["razorpay_order_id"] = razorpay_order["id"]
        transaction["payment_gateway_id"] = razorpay_order["id"]
        
        # Save transaction to database
        await db.transactions.insert_one(transaction)
        
        # Return order details for frontend
        return PaymentOrderResponse(
            order_id=razorpay_order["id"],
            amount=amount_in_paise,
            currency=order_request.currency,
            key_id=os.getenv("RAZORPAY_KEY_ID"),
            transaction_id=transaction_id
        )
    
    except Exception as e:
        # Mark transaction as failed
        transaction["status"] = TransactionStatus.FAILED
        transaction["notes"] = f"Order creation failed: {str(e)}"
        await db.transactions.insert_one(transaction)
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create payment order: {str(e)}"
        )


@api_router.post("/payment/verify", response_model=PaymentVerificationResponse)
@limiter.limit(RATE_LIMIT_DEFAULT)
async def verify_payment(
    request: Request,
    verification: PaymentVerificationRequest,
    current_user: dict = Depends(get_current_user)
):
    """Verify Razorpay payment and update credits/subscription"""
    
    # Find the transaction
    transaction = await db.transactions.find_one({
        "id": verification.transaction_id,
        "user_id": current_user["id"]
    })
    
    if not transaction:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Transaction not found"
        )
    
    # Check if already processed
    if transaction["status"] == TransactionStatus.COMPLETED:
        subscription = await db.subscriptions.find_one({"user_id": current_user["id"]})
        return PaymentVerificationResponse(
            success=True,
            transaction_id=verification.transaction_id,
            credits_added=transaction.get("credits_added"),
            new_balance=subscription.get("credits"),
            plan_upgraded=transaction.get("plan_upgraded_to"),
            message="Payment already processed"
        )
    
    # Verify payment signature
    try:
        is_valid = verify_razorpay_payment(
            verification.razorpay_order_id,
            verification.razorpay_payment_id,
            verification.razorpay_signature
        )
        
        if not is_valid:
            # Mark as failed
            await db.transactions.update_one(
                {"id": verification.transaction_id},
                {
                    "$set": {
                        "status": TransactionStatus.FAILED,
                        "notes": "Invalid payment signature"
                    }
                }
            )
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid payment signature"
            )
        
        # Payment verified successfully - update transaction
        now = datetime.now(timezone.utc)
        await db.transactions.update_one(
            {"id": verification.transaction_id},
            {
                "$set": {
                    "razorpay_payment_id": verification.razorpay_payment_id,
                    "razorpay_signature": verification.razorpay_signature,
                    "status": TransactionStatus.COMPLETED,
                    "completed_at": now
                }
            }
        )
        
        # Update credits or subscription atomically
        credits_added = transaction.get("credits_added")
        plan_upgraded = transaction.get("plan_upgraded_to")
        
        if credits_added:
            # Add credits to user's account
            result = await db.subscriptions.find_one_and_update(
                {"user_id": current_user["id"]},
                {"$inc": {"credits": credits_added}, "$set": {"updated_at": now}},
                return_document=True
            )
            new_balance = result["credits"]
        else:
            new_balance = None
        
        if plan_upgraded:
            # Upgrade subscription plan
            plan_config = PLAN_CONFIGS.get(plan_upgraded)
            if plan_config:
                await db.subscriptions.update_one(
                    {"user_id": current_user["id"]},
                    {
                        "$set": {
                            "subscription_tier": plan_upgraded,
                            "monthly_invoice_limit": plan_config.monthly_invoice_limit,
                            "subscription_start_date": now,
                            "subscription_end_date": now + timedelta(days=30),
                            "updated_at": now
                        },
                        "$inc": {"credits": plan_config.initial_credits}
                    }
                )
        
        return PaymentVerificationResponse(
            success=True,
            transaction_id=verification.transaction_id,
            credits_added=credits_added,
            new_balance=new_balance,
            plan_upgraded=plan_upgraded,
            message="Payment verified successfully"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        # Mark as failed
        await db.transactions.update_one(
            {"id": verification.transaction_id},
            {
                "$set": {
                    "status": TransactionStatus.FAILED,
                    "notes": f"Verification error: {str(e)}"
                }
            }
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Payment verification failed: {str(e)}"
        )


@api_router.post("/payment/webhook")
async def razorpay_webhook(request: Request):
    """Handle Razorpay webhook events"""
    
    # Get raw body and signature
    body = await request.body()
    signature = request.headers.get("X-Razorpay-Signature", "")
    
    # Verify webhook signature
    if not verify_razorpay_webhook(body, signature):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid webhook signature"
        )
    
    # Parse webhook payload
    try:
        payload = json.loads(body)
        event = payload.get("event")
        payment_entity = payload.get("payload", {}).get("payment", {}).get("entity", {})
        
        # Handle different event types
        if event == "payment.captured":
            # Payment successful
            order_id = payment_entity.get("order_id")
            payment_id = payment_entity.get("id")
            
            # Find transaction by order_id
            transaction = await db.transactions.find_one({"razorpay_order_id": order_id})
            
            if transaction and transaction["status"] == TransactionStatus.PENDING:
                # Update transaction status
                now = datetime.now(timezone.utc)
                await db.transactions.update_one(
                    {"id": transaction["id"]},
                    {
                        "$set": {
                            "razorpay_payment_id": payment_id,
                            "status": TransactionStatus.COMPLETED,
                            "completed_at": now
                        }
                    }
                )
                
                # Add credits or upgrade plan
                if transaction.get("credits_added"):
                    await db.subscriptions.update_one(
                        {"user_id": transaction["user_id"]},
                        {
                            "$inc": {"credits": transaction["credits_added"]},
                            "$set": {"updated_at": now}
                        }
                    )
                
                if transaction.get("plan_upgraded_to"):
                    plan = transaction["plan_upgraded_to"]
                    plan_config = PLAN_CONFIGS.get(plan)
                    if plan_config:
                        await db.subscriptions.update_one(
                            {"user_id": transaction["user_id"]},
                            {
                                "$set": {
                                    "subscription_tier": plan,
                                    "monthly_invoice_limit": plan_config.monthly_invoice_limit,
                                    "subscription_start_date": now,
                                    "subscription_end_date": now + timedelta(days=30),
                                    "updated_at": now
                                },
                                "$inc": {"credits": plan_config.initial_credits}
                            }
                        )
        
        elif event == "payment.failed":
            # Payment failed
            order_id = payment_entity.get("order_id")
            
            # Update transaction status
            await db.transactions.update_one(
                {"razorpay_order_id": order_id},
                {
                    "$set": {
                        "status": TransactionStatus.FAILED,
                        "notes": f"Payment failed: {payment_entity.get('error_description', 'Unknown error')}"
                    }
                }
            )
        
        return {"status": "ok"}
    
    except Exception as e:
        logging.error(f"Webhook processing error: {str(e)}")
        return {"status": "error", "message": str(e)}


@api_router.get("/payment/gateways")
async def get_payment_gateways():
    """Get list of configured payment gateways"""
    gateways = get_configured_gateways()
    return {
        "available_gateways": gateways,
        "primary": "razorpay" if "razorpay" in gateways else None
    }

# ─── Health ───────────────────────────────────────────────────────────────────

@api_router.get("/health")
async def health():
    return {"status": "ok", "service": "InvoiceAI Backend"}

app.include_router(api_router)
app.add_middleware(MaxRequestSizeMiddleware, max_body_size=MAX_UPLOAD_BYTES)
app.add_middleware(UserIdContextMiddleware)
app.add_middleware(SlowAPIMiddleware)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=CORS_ALLOW_ORIGINS,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    mongo_client.close()

# ═══════════════════════════════════════════════════════════════════════════════
# SERVER STARTUP
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")
